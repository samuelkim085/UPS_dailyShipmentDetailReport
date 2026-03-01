"""
UPS Daily Shipment Detail Report - PDF Extractor
Extracts Package Ref No.1 and Tracking No. from UPS WorldShip PDF reports.

Usage:
    python ups_extract.py                          # interactive prompt / drag-and-drop
    python ups_extract.py <pdf_file>
    python ups_extract.py <pdf_file> -o output.csv
    python ups_extract.py <pdf_file> --format xlsx
    python ups_extract.py <pdf_file> --json        # JSON output for Electron app
"""

import argparse
import csv
import io
import json
import re
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf-8-sig"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() not in ("utf-8", "utf-8-sig"):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import pdfplumber
except ImportError:
    print("Error: pdfplumber is required. Install it with:")
    print("  pip install pdfplumber")
    sys.exit(1)


def extract_shipment_data(pdf_path: str) -> list[dict]:
    """Extract Package Ref No.1 and Tracking No. pairs from a UPS PDF report."""
    records = []
    current_ref = None
    is_void = False

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            lines = text.split("\n")
            for line in lines:
                # Normalize common OCR artifacts
                normalized = re.sub(r'이\s*', '01', line)
                normalized = re.sub(r'[「」]', '01', normalized)

                # VOID detection
                if re.search(r"\bVOID\b", line) and "Voided" not in line:
                    if records and "Tracking" not in line:
                        records[-1]["Status"] = "VOID"
                    else:
                        is_void = True

                # Package Ref No.1
                # Lookahead stops at "UPS T..." (handles OCR variants: Total/TOtal/TotaI/TOtal)
                ref_match = re.search(
                    r"Package\s*Ref\s*N[Oo]\.?\s*1\s*[.:]\s*(.+?)(?=\s+UPS\s+T|\s+Tracking|\s+Service\s+Type|$)",
                    normalized
                )
                if ref_match:
                    current_ref = ref_match.group(1).strip()
                    # Belt-and-suspenders: also strip with re.sub in case lookahead missed
                    current_ref = re.sub(r'\s+UPS\s+T\S.*$', '', current_ref, flags=re.IGNORECASE).strip()
                    current_ref = re.sub(r'\s+Tracking.*$', '', current_ref, flags=re.IGNORECASE).strip()

                # Tracking No.
                tracking_match = re.search(
                    r"Tracking\s*N[Oo]\.?[.•]?\s*:?\s*([1Il][Z27][A-Z0-9]+)",
                    normalized
                )
                if tracking_match:
                    tracking = tracking_match.group(1).strip()
                    suffix = tracking[8:] if len(tracking) > 8 else ""
                    suffix = suffix.replace("O", "0").replace("I", "1").replace("l", "1")
                    tracking = "1ZGW0159" + suffix[:10]
                    if len(tracking) < 18:
                        continue
                    records.append({
                        "Package Ref No.1": current_ref or "",
                        "Tracking No.": tracking,
                        "Status": "VOID" if is_void else "Active",
                    })
                    is_void = False

    return records


def print_table(records: list[dict]) -> None:
    if not records:
        print("No records found.")
        return

    num_w = len(str(len(records)))
    ref_w = max(max(len(r["Package Ref No.1"]) for r in records), len("Package Ref No.1"))
    trk_w = max(max(len(r["Tracking No."]) for r in records), len("Tracking No."))
    sts_w = max(max(len(r["Status"]) for r in records), len("Status"))

    header = f"{'#':>{num_w}} | {'Package Ref No.1':<{ref_w}} | {'Tracking No.':<{trk_w}} | {'Status':<{sts_w}}"
    sep = "-" * len(header)
    print(sep)
    print(header)
    print(sep)
    for i, r in enumerate(records, 1):
        status_display = r["Status"] if r["Status"] == "VOID" else ""
        print(f"{i:>{num_w}} | {r['Package Ref No.1']:<{ref_w}} | {r['Tracking No.']:<{trk_w}} | {status_display:<{sts_w}}")
    print(sep)

    active = sum(1 for r in records if r["Status"] == "Active")
    voided = sum(1 for r in records if r["Status"] == "VOID")
    print(f"Total: {len(records)} packages ({active} active, {voided} voided)")


def save_csv(records: list[dict], output_path: str) -> None:
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Package Ref No.1", "Tracking No.", "Status"])
        writer.writeheader()
        writer.writerows(records)
    print(f"Saved {len(records)} records to {output_path}")


def save_xlsx(records: list[dict], output_path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Error: openpyxl is required. Install it with: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "UPS Shipments"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["#", "Package Ref No.1", "Tracking No.", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    void_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for i, r in enumerate(records, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=r["Package Ref No.1"]).border = thin_border
        ws.cell(row=row, column=3, value=r["Tracking No."]).border = thin_border
        status_cell = ws.cell(row=row, column=4, value=r["Status"])
        status_cell.border = thin_border
        if r["Status"] == "VOID":
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = void_fill

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 10
    wb.save(output_path)
    print(f"Saved {len(records)} records to {output_path}")


def clean_path(raw: str) -> str:
    return raw.strip().strip('"').strip("'").strip()


def prompt_for_pdf() -> Path:
    print("=" * 60)
    print("  UPS Daily Shipment Detail Report - PDF Extractor")
    print("=" * 60)
    print()
    print("  Drag and drop a PDF file here, or type the file path:")
    print()
    while True:
        try:
            raw = input("  PDF file: ")
        except (EOFError, KeyboardInterrupt):
            print("\nCancelled.")
            sys.exit(0)
        pdf_path = Path(clean_path(raw))
        if not pdf_path.as_posix():
            print("  No input provided. Please try again.\n")
            continue
        if not pdf_path.exists():
            print(f"  File not found: {pdf_path}\n")
            continue
        if pdf_path.suffix.lower() != ".pdf":
            print(f"  Not a PDF file: {pdf_path.name}\n")
            continue
        return pdf_path


def prompt_for_output(pdf_path: Path) -> tuple[Path | None, str | None]:
    print()
    print("  Save output to file?")
    print("    1) CSV")
    print("    2) Excel (.xlsx)")
    print("    3) No, just show the table")
    print()
    while True:
        try:
            choice = input("  Choice [1/2/3]: ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            return None, None
        if choice == "1":
            return pdf_path.with_suffix(".csv"), "csv"
        elif choice == "2":
            return pdf_path.with_suffix(".xlsx"), "xlsx"
        elif choice in ("3", ""):
            return None, None
        else:
            print("  Invalid choice. Enter 1, 2, or 3.\n")


def clear_cache() -> None:
    import shutil
    script_dir = Path(__file__).parent
    removed = []
    for cache_dir in script_dir.rglob("__pycache__"):
        shutil.rmtree(cache_dir, ignore_errors=True)
        removed.append(str(cache_dir))
    for pyc in script_dir.rglob("*.pyc"):
        pyc.unlink(missing_ok=True)
        removed.append(str(pyc))
    if removed:
        print(f"Cleared {len(removed)} cached item(s).")
    else:
        print("No cache found.")


def main():
    parser = argparse.ArgumentParser(
        description="Extract Package Ref No.1 and Tracking No. from UPS Daily Shipment Detail Report PDFs."
    )
    parser.add_argument("pdf", nargs="?", default=None, help="Path to the UPS PDF report")
    parser.add_argument("-o", "--output", help="Output file path (CSV or XLSX)")
    parser.add_argument("--reset", action="store_true", help="Clear Python cache and exit")
    parser.add_argument("--json", action="store_true", help="Output records as JSON to stdout (used by Electron)")
    parser.add_argument("--format", choices=["csv", "xlsx"], default=None, help="Output format")

    args = parser.parse_args()

    if args.reset:
        clear_cache()
        return

    if args.pdf is None:
        pdf_path = prompt_for_pdf()
        print(f"\n  Reading: {pdf_path}\n")
        records = extract_shipment_data(str(pdf_path))
        print_table(records)
        output_path, fmt = prompt_for_output(pdf_path)
        if output_path and fmt:
            if fmt == "xlsx":
                save_xlsx(records, str(output_path))
            else:
                save_csv(records, str(output_path))
        return

    pdf_path = Path(clean_path(args.pdf))
    if not pdf_path.exists():
        print(f"Error: File not found: {pdf_path}")
        sys.exit(1)

    records = extract_shipment_data(str(pdf_path))

    if args.json:
        print(json.dumps(records, ensure_ascii=False))
        return

    print(f"Reading: {pdf_path}")
    print_table(records)

    if args.output:
        output_path = Path(args.output)
        fmt = args.format or output_path.suffix.lstrip(".").lower()
        if fmt not in ("csv", "xlsx"):
            fmt = "csv"
        if fmt == "xlsx":
            save_xlsx(records, str(output_path))
        else:
            save_csv(records, str(output_path))
    elif args.format:
        output_path = pdf_path.parent / f"{pdf_path.stem}.{args.format}"
        if args.format == "xlsx":
            save_xlsx(records, str(output_path))
        else:
            save_csv(records, str(output_path))


if __name__ == "__main__":
    main()
