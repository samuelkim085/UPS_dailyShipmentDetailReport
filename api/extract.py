import csv
import io
import os
import sys
import tempfile

from flask import Flask, request, jsonify, render_template, send_file

# Add parent directory to path so we can import ups_extract
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from ups_extract import extract_shipment_data

app = Flask(
    __name__,
    template_folder=os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "templates"),
)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/extract", methods=["POST"])
def extract():
    if "pdf" not in request.files:
        return jsonify({"error": "No PDF file uploaded"}), 400

    pdf_file = request.files["pdf"]
    if not pdf_file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "File must be a PDF"}), 400

    # Save to temp file, extract, then clean up
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        pdf_file.save(tmp.name)
        tmp_path = tmp.name

    try:
        records = extract_shipment_data(tmp_path)
        return jsonify({"records": records})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        os.unlink(tmp_path)


@app.route("/api/download", methods=["POST"])
def download():
    data = request.get_json()
    records = data.get("records", [])
    fmt = data.get("format", "csv")

    if fmt == "xlsx":
        return _build_xlsx(records)
    return _build_csv(records)


def _build_csv(records):
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["Package Ref No.1", "Tracking No.", "Status"])
    writer.writeheader()
    for r in records:
        writer.writerow({
            "Package Ref No.1": r.get("Package Ref No.1", ""),
            "Tracking No.": r.get("Tracking No.", ""),
            "Status": r.get("Status", ""),
        })
    buf.seek(0)
    return send_file(
        io.BytesIO(buf.getvalue().encode("utf-8")),
        mimetype="text/csv",
        as_attachment=True,
        download_name="shipments.csv",
    )


def _build_xlsx(records):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "UPS Shipments"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["#", "Package Ref No.1", "Tracking No.", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    void_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for i, r in enumerate(records, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=r.get("Package Ref No.1", "")).border = thin_border
        ws.cell(row=row, column=3, value=r.get("Tracking No.", "")).border = thin_border
        status_cell = ws.cell(row=row, column=4, value=r.get("Status", ""))
        status_cell.border = thin_border
        if r.get("Status") == "VOID":
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = void_fill

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="shipments.xlsx",
    )
